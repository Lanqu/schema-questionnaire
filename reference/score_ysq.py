import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

answers = {
    1: 4,
    2: 4,
    3: 4,
    4: 4,
    5: 2,
    6: 1,
    7: 2,
    8: 4,
    9: 2,
    10: 2,
    11: 3,
    12: 2,
    13: 4,
    14: 2,
    15: 3,
    16: 2,
    17: 4,
    18: 2,
    19: 1,
    20: 4,
    21: 3,
    22: 3,
    23: 1,
    24: 2,
    25: 3,
    26: 4,
    27: 3,
    28: 5,
    29: 4,
    30: 4,
    31: 4,
    32: 2,
    33: 1,
    34: 4,
    35: 3,
    36: 3,
    37: 1,
    38: 3,
    39: 3,
    40: 4,
    41: 1,
    42: 2,
    43: 2,
    44: 4,
    45: 4,
    46: 2,
    47: 3,
    48: 3,
    49: 5,
    50: 4,
    51: 4,
    52: 2,
    53: 4,
    54: 4,
    55: 3,
    56: 4,
    57: 5,
    58: 2,
    59: 2,
    60: 2,
    61: 2,
    62: 4,
    63: 3,
    64: 1,
    65: 4,
    66: 3,
    67: 4,
    68: 2,
    69: 4,
    70: 3,
    71: 5,
    72: 4,
    73: 4,
    74: 4,
    75: 5,
    76: 3,
    77: 3,
    78: 2,
    79: 3,
    80: 3,
    81: 3,
    82: 4,
    83: 6,
    84: 4,
    85: 4,
    86: 3,
    87: 4,
    88: 3,
    89: 3,
    90: 1,
}

# Schema mapping: each schema has 5 questions cycling through 18 schemas
schemas = [
    ("ЭД", "Эмоциональная депривация", "РСО", [1, 19, 37, 55, 73]),
    ("ПН", "Покинутость / нестабильность", "РСО", [2, 20, 38, 56, 74]),
    ("НН", "Недоверие / насилие", "РСО", [3, 21, 39, 57, 75]),
    ("СИО", "Социальная изоляция / отчуждение", "РСО", [4, 22, 40, 58, 76]),
    ("ДС", "Дефективность / стыд", "РСО", [5, 23, 41, 59, 77]),
    ("НО", "Неуспешность / обреченность на неудачу", "НЛАНД", [6, 24, 42, 60, 78]),
    ("ЗН", "Зависимость / некомпетентность", "НЛАНД", [7, 25, 43, 61, 79]),
    ("У", "Уязвимость", "НЛАНД", [8, 26, 44, 62, 80]),
    ("СДНЯ", "Слитность с другими / неразвитое Я", "НЛАНД", [9, 27, 45, 63, 81]),
    ("ПП", "Покорность / подчинение", "НД", [10, 28, 46, 64, 82]),
    ("С", "Самопожертвование", "НД", [11, 29, 47, 65, 83]),
    ("ПЭ", "Подавление эмоций", "СбП", [12, 30, 48, 66, 84]),
    ("ЖС", "Жёсткие стандарты / завышенные требования", "СбП", [13, 31, 49, 67, 85]),
    ("ПГ", "Привилегированность / грандиозность", "НГ", [14, 32, 50, 68, 86]),
    ("НС", "Недостаток самоконтроля / самодисциплины", "НГ", [15, 33, 51, 69, 87]),
    ("ПО", "Поиск одобрения", "НД", [16, 34, 52, 70, 88]),
    ("НП", "Негативизм / пессимизм", "СбП", [17, 35, 53, 71, 89]),
    ("ПК", "Пунитивность / карательность", "СбП", [18, 36, 54, 72, 90]),
]

# Domain groupings
domains = {
    "РСО": ("Разобщенность и отвержение", ["ЭД", "ПН", "НН", "СИО", "ДС"]),
    "НЛАНД": (
        "Нарушение личностной автономии и непризнание достижений",
        ["НО", "ЗН", "У", "СДНЯ"],
    ),
    "НГ": ("Нарушенные границы", ["ПГ", "НС"]),
    "НД": ("Направленность на других", ["ПП", "С", "ПО"]),
    "СбП": ("Сверхбдительность и подавление", ["ПЭ", "ЖС", "НП", "ПК"]),
}

print("=" * 80)
print("YSQ-S3R RESULTS")
print("=" * 80)

results = {}
for code, name, domain, qs in schemas:
    scores = [answers[q] for q in qs]
    total = sum(scores)
    mean = total / 5
    pct = (total - 5) / 25 * 100

    if pct <= 20:
        level = "низкий"
    elif pct <= 40:
        level = "пониженный"
    elif pct <= 60:
        level = "умеренный"
    elif pct <= 80:
        level = "повышенный"
    else:
        level = "высокий"

    high_count = sum(1 for s in scores if s >= 5)
    results[code] = {
        "name": name,
        "domain": domain,
        "scores": scores,
        "mean": mean,
        "total": total,
        "pct": pct,
        "level": level,
        "high_count": high_count,
    }

# Sort by score descending for display
sorted_schemas = sorted(results.items(), key=lambda x: x[1]["mean"], reverse=True)

# Print by domain
for domain_code, (domain_name, schema_codes) in domains.items():
    domain_schemas = [(c, results[c]) for c in schema_codes]
    domain_mean = sum(r["mean"] for _, r in domain_schemas) / len(domain_schemas)
    domain_high = sum(r["high_count"] for _, r in domain_schemas)
    print(f"\n{'─' * 80}")
    print(
        f"  {domain_name} ({domain_code}) — Domain avg: {domain_mean:.1f}, High (5-6): {domain_high}"
    )
    print(f"{'─' * 80}")
    for code, r in domain_schemas:
        bar = "█" * int(r["mean"]) + "░" * (6 - int(r["mean"]))
        print(f"  {code:5s} {r['name'][:50]:50s} {r['mean']:.1f}  {bar}  {r['level']}")
        print(
            f"         Items: {r['scores']}  Sum: {r['total']}  %: {r['pct']:.0f}%  High(5-6): {r['high_count']}"
        )

print(f"\n{'=' * 80}")
print("TOP SCHEMAS (sorted by mean score)")
print(f"{'=' * 80}")
for code, r in sorted_schemas:
    marker = " *** " if r["mean"] >= 4.0 else "     "
    print(f"{marker}{code:5s} {r['name'][:45]:45s} {r['mean']:.1f}  ({r['level']})")

# Write scores into the Excel file
wb = openpyxl.load_workbook("F:/personal/psyco/YSQ-RUS (7).xlsx")
ws = wb["Опросник"]

# Find the score column (Q = column 17, 1-indexed -> 17 in openpyxl)
# Questions are at specific rows - need to map question number to row
# From exploration: q1 at row 22, q2 at row 24, etc. (every 2 rows)
# But there are scale legend breaks. Let me find exact rows.
q_rows = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    val = row[1].value  # col B
    if val and isinstance(val, str):
        import re

        m = re.match(r"^(\d+)\.\s+", val.strip())
        if m:
            q_rows[int(m.group(1))] = row[0].row

print(f"\nFound {len(q_rows)} question rows in Excel")

# Write scores to column R (18th column, index 17 in 0-based, col 18 in openpyxl)
# The score column is Q (17th column) based on the formula references
for qnum, score in answers.items():
    if qnum in q_rows:
        row_num = q_rows[qnum]
        ws.cell(row=row_num, column=17, value=score)  # Column Q

wb.save("F:/personal/psyco/YSQ-RUS (7).xlsx")
print("Scores written to YSQ-RUS (7).xlsx")
